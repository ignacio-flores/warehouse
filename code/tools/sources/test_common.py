import pathlib
import shutil
import sys
import tempfile
import unittest
import zipfile
from unittest import mock

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - exercised in dependency failures
    load_workbook = None
    OPENPYXL_IMPORT_ERROR = exc
else:
    OPENPYXL_IMPORT_ERROR = None

SOURCE_DIR = pathlib.Path(__file__).resolve().parent
REPO_ROOT = SOURCE_DIR.parents[2]
SOURCE_WORKBOOK = REPO_ROOT / "handmade_tables/dictionary.xlsx"

if str(SOURCE_DIR) not in sys.path:
    sys.path.insert(0, str(SOURCE_DIR))

import common


class UrlNormalizationTests(unittest.TestCase):
    def test_normalize_url_preserves_fragment_identity(self):
        first = common.normalize_url(
            "https://research.ibfd.org/#/doc?url=/linkresolver/static/gthb_bj_2020-06-30_s_5."
        )
        second = common.normalize_url(
            "https://research.ibfd.org/#/doc?url=/linkresolver/static/gthb_me_2025-05-10_s_5."
        )

        self.assertNotEqual(first, second)
        self.assertIn("#/doc?url=/linkresolver/static/gthb_bj_2020-06-30_s_5.", first)

    def test_normalize_url_keeps_existing_path_query_behavior(self):
        self.assertEqual(
            common.normalize_url(" HTTPS://Example.COM//foo///bar/?x=1 "),
            "https://example.com/foo/bar?x=1",
        )


class BibTaxonomyTests(unittest.TestCase):
    def test_migrates_section_to_canonical_keyword_and_backfills_blank(self):
        result = common.migrate_keywords_for_data_source_section("Inheritance Trends", "")

        self.assertTrue(result["backfilled"])
        self.assertTrue(result["changed"])
        self.assertEqual(result["after_keywords"], "Data Sources: Inheritance Trends")
        self.assertEqual(result["after_data_source_keywords"], ["Data Sources: Inheritance Trends"])

    def test_migration_replaces_stale_data_source_keyword_and_preserves_research_topic(self):
        result = common.migrate_keywords_for_data_source_section(
            "Taxes on Wealth",
            "Data Sources: Estate Inheritance and Gift Taxes,Estate Inheritance and Gift Taxes",
        )

        self.assertEqual(
            result["after_keywords"],
            "Data Sources: Taxes on Wealth,Estate Inheritance and Gift Taxes",
        )

    def test_bib_rendering_never_emits_section_field(self):
        rendered_record = common.render_bib_entry(
            "Example2026",
            {
                "bib": {
                    "entry_type": "misc",
                    "title": "Example",
                    "author": "Example",
                    "year": "2026",
                    "keywords": "Data Sources: Wealth Topography",
                    "extra_fields": {"section": "Wealth Topography", "url_file": "example.csv"},
                }
            },
        )
        rendered_parsed = common.render_parsed_bib_entry(
            "Parsed2026",
            {
                "entry_type": "misc",
                "fields": {
                    "title": "Parsed",
                    "section": "Wealth Topography",
                    "keywords": "Data Sources: Wealth Topography",
                },
            },
        )

        self.assertNotIn("section =", rendered_record)
        self.assertNotIn("section =", rendered_parsed)
        self.assertIn("url_file =", rendered_record)

    def test_digital_library_merge_unions_keywords_prefers_registry_urls_and_reports_conflicts(self):
        records = [
            {
                "id": "src-shared",
                "is_data_source": True,
                "section": "Taxes on Wealth",
                "source": "Shared2026",
                "citekey": "Shared2026",
                "bib": {
                    "entry_type": "misc",
                    "title": "Registry Title",
                    "author": "Registry Author",
                    "year": "2026",
                    "url": "https://data.example/source",
                    "keywords": "Data Sources: Taxes on Wealth,Wealth Taxation",
                    "extra_fields": {
                        "data_file": "https://data.example/file.csv",
                        "section": "Hidden",
                    },
                },
            },
            {
                "id": "src-research",
                "is_data_source": False,
                "section": "",
                "source": "Shared2026Research",
                "citekey": "Shared2026",
                "bib": {
                    "entry_type": "article",
                    "title": "Wealth Title",
                    "author": "Wealth Author",
                    "year": "2026",
                    "url": "https://wealth.example/paper",
                    "keywords": "Intergenerational Wealth",
                    "extra_fields": {
                        "section": "Legacy",
                        "replication_package": "https://wealth.example/replication",
                    },
                },
            },
        ]

        entries, report = common.build_digital_library_entries(records)
        fields = entries["Shared2026"]["fields"]

        self.assertEqual(fields["title"], "Registry Title")
        self.assertEqual(fields["url"], "https://data.example/source")
        self.assertEqual(
            fields["keywords"],
            "Data Sources: Taxes on Wealth,Wealth Taxation,Intergenerational Wealth",
        )
        self.assertEqual(fields["data_file"], "https://data.example/file.csv")
        self.assertEqual(fields["replication_package"], "https://wealth.example/replication")
        self.assertNotIn("section", fields)
        self.assertIn("url_2", fields)
        self.assertEqual({row["field"] for row in report["bibliographic_conflicts"]}, {"title", "author"})

    def test_digital_library_duplicate_citekey_unions_cross_category_keywords(self):
        records = [
            {
                "id": "src-topography",
                "section": "Wealth Topography",
                "source": "SharedData2026",
                "citekey": "SharedData2026",
                "bib": {
                    "entry_type": "misc",
                    "title": "Shared Dataset",
                    "author": "Data Office",
                    "year": "2026",
                    "keywords": "Data Sources: Wealth Topography",
                },
            },
            {
                "id": "src-inequality",
                "section": "Wealth Inequality",
                "source": "SharedData2026",
                "citekey": "SharedData2026",
                "bib": {
                    "entry_type": "misc",
                    "title": "Shared Dataset",
                    "author": "Data Office",
                    "year": "2026",
                    "keywords": "Data Sources: Wealth Inequality,Distributional Accounts",
                },
            },
        ]

        entries, report = common.build_digital_library_entries(records)

        self.assertEqual(
            entries["SharedData2026"]["fields"]["keywords"],
            "Data Sources: Wealth Topography,Data Sources: Wealth Inequality,Distributional Accounts",
        )
        self.assertEqual([row["citekey"] for row in report["duplicate_citekeys"]], ["SharedData2026"])
        self.assertEqual([row["citekey"] for row in report["multi_data_source_keyword_exports"]], ["SharedData2026"])


class ExcelHelperTests(unittest.TestCase):
    def test_sanitize_excel_string_removes_invalid_xml_chars(self):
        raw = "Alpha" + chr(0) + "Beta" + chr(0xD800) + "\nGamma"
        self.assertEqual(common.sanitize_excel_string(raw), "AlphaBeta\nGamma")

    def test_sanitize_sheet_name_strips_invalid_chars_and_truncates(self):
        sheet_name = common.sanitize_sheet_name(" 'Bad:/\\[]*? sheet name that is far too long' ")
        self.assertLessEqual(len(sheet_name), 31)
        self.assertNotRegex(sheet_name, r"[\[\]:*?/\\]")
        self.assertFalse(sheet_name.startswith("'"))
        self.assertFalse(sheet_name.endswith("'"))
        self.assertTrue(sheet_name.startswith("Bad"))

    def test_sanitize_excel_value_replaces_non_finite_numbers(self):
        self.assertEqual(common.sanitize_excel_value(float("nan")), "")
        self.assertEqual(common.sanitize_excel_value(float("inf")), "")
        self.assertEqual(common.sanitize_excel_value(float("-inf")), "")
        self.assertEqual(common.sanitize_excel_value(12.5), "12.5")


class DictionaryWorkbookWriteTests(unittest.TestCase):
    def _assert_openpyxl_available(self):
        if load_workbook is None:
            self.fail(
                "openpyxl test dependency is missing. "
                "Install it with: python3 -m pip install openpyxl"
            )

    def _assert_valid_xlsx(self, workbook_path: pathlib.Path):
        self.assertTrue(workbook_path.exists(), f"Workbook missing: {workbook_path}")
        with zipfile.ZipFile(workbook_path, "r") as archive:
            self.assertIsNone(archive.testzip())
        self._assert_openpyxl_available()
        workbook = load_workbook(workbook_path, read_only=True)
        try:
            self.assertIn("Sources", workbook.sheetnames)
        finally:
            workbook.close()

    def _sample_row(self, **overrides):
        row = {header: "" for header in common.SOURCES_HEADERS}
        row.update(
            {
                "Section": "Wealth",
                "Legend": "Example legend",
                "Source": "Example Source",
                "Link": "https://example.com/source",
                "Citekey": "ExampleSource2026",
            }
        )
        row.update(overrides)
        return row

    def test_invalid_temp_workbook_does_not_replace_existing_output(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = pathlib.Path(tmpdir)
            template = tmpdir_path / "template.xlsx"
            output = tmpdir_path / "dictionary.xlsx"
            shutil.copy2(SOURCE_WORKBOOK, template)
            shutil.copy2(SOURCE_WORKBOOK, output)
            before = output.read_bytes()

            with mock.patch.object(common, "workbook_to_xlsx_bytes", return_value=b"not a zip workbook"):
                with self.assertRaises(RuntimeError):
                    common.write_sources_sheet(template, output, [])

            self.assertEqual(output.read_bytes(), before)
            self.assertEqual(
                sorted(path.name for path in tmpdir_path.iterdir()),
                ["dictionary.xlsx", "template.xlsx"],
            )

    def test_validate_xlsx_file_rejects_corrupt_content(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            bad_path = pathlib.Path(tmpdir) / "broken.xlsx"
            bad_path.write_bytes(b"not a zip file")
            with self.assertRaises(RuntimeError):
                common.validate_xlsx_file(bad_path)

    def test_write_sources_sheet_sanitizes_values_and_validates_output(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = pathlib.Path(tmpdir)
            template = tmpdir_path / "template.xlsx"
            output = tmpdir_path / "dictionary.xlsx"
            shutil.copy2(SOURCE_WORKBOOK, template)

            rows = [
                self._sample_row(
                    Legend="Alpha" + chr(0) + "Beta",
                    Metadata=float("nan"),
                    Metadatalink=float("inf"),
                    TAreply=float("-inf"),
                    QcommentsforTA="Needs" + chr(0x1F) + " review",
                )
            ]

            common.write_sources_sheet(template, output, rows)

            common.validate_xlsx_file(output)
            common.validate_xlsx_file(output.read_bytes())
            self._assert_valid_xlsx(output)

            written_rows = common.read_sources_sheet(output)
            self.assertEqual(len(written_rows), 1)
            self.assertEqual(written_rows[0]["Legend"], "AlphaBeta")
            self.assertEqual(written_rows[0]["Metadata"], "")
            self.assertEqual(written_rows[0]["Metadatalink"], "")
            self.assertEqual(written_rows[0]["TAreply"], "")
            self.assertEqual(written_rows[0]["QcommentsforTA"], "Needs review")

    def test_write_sources_sheet_supports_in_place_rewrite(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = pathlib.Path(tmpdir) / "dictionary.xlsx"
            shutil.copy2(SOURCE_WORKBOOK, workbook_path)

            rows = [self._sample_row(Source="In Place Source", Citekey="InPlace2026")]
            common.write_sources_sheet(workbook_path, workbook_path, rows)

            common.validate_xlsx_file(workbook_path)
            self._assert_valid_xlsx(workbook_path)
            written_rows = common.read_sources_sheet(workbook_path)
            self.assertEqual(len(written_rows), 1)
            self.assertEqual(written_rows[0]["Source"], "In Place Source")
            self.assertEqual(written_rows[0]["Citekey"], "InPlace2026")


if __name__ == "__main__":
    unittest.main()
