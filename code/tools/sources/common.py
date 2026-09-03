#!/usr/bin/env python3
"""Common utilities for source registry tooling."""

import io
import json
import math
import re
import tempfile
import zipfile
from collections import OrderedDict
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple
from urllib.parse import urlparse
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape

from source_paths import (
    DEFAULT_BOTH_BIB_PATH,
    DEFAULT_DATA_BIB_PATH,
    DEFAULT_DIGITAL_BIB_PATH,
    DEFAULT_DICTIONARY_PATH,
    DEFAULT_WEALTH_BIB_PATH,
    DEFAULT_WEALTH_CHANGE_LOG_PATH,
)

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
EXPECTED_SOURCES_SHEET_NAME = "Sources"
REQUIRED_XLSX_MEMBERS = ("xl/workbook.xml", "xl/_rels/workbook.xml.rels")
INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]:*?/\\]")

SOURCES_HEADERS = [
    "Section",
    "AggSource",
    "Legend",
    "Source",
    "Data_Type",
    "Link",
    "Ref_link",
    "Citekey",
    "Inclusion_in_Warehouse",
    "Multigeo_Reference",
    "Metadata",
    "Metadatalink",
    "QcommentsforTA",
    "TAreply",
    "TAcomments",
    "ARJcomments",
    "ARJreplies",
    "SeeAggSourcelisthere",
]

CANONICAL_KEYS = [
    "id",
    "section",
    "aggsource",
    "legend",
    "source",
    "data_type",
    "link",
    "ref_link",
    "citekey",
    "inclusion_in_warehouse",
    "multigeo_reference",
    "metadata",
    "metadatalink",
    "qcommentsforta",
    "tareply",
    "tacomments",
    "arjcomments",
    "arjreplies",
    "seeaggsourcelisthere",
    "bib",
    "created_at",
    "updated_at",
]

BIB_FIELD_ORDER = [
    "title",
    "author",
    "year",
    "month",
    "journal",
    "booktitle",
    "volume",
    "number",
    "pages",
    "institution",
    "publisher",
    "doi",
    "url",
    "urldate",
    "abstract",
    "keywords",
    "note",
]

DATA_SOURCE_KEYWORD_PREFIX = "Data Sources: "
CANONICAL_DATA_SOURCE_KEYWORDS = [
    "Data Sources: Wealth Topography",
    "Data Sources: Wealth Inequality",
    "Data Sources: Taxes on Wealth",
    "Data Sources: Inheritance Trends",
    "Data Sources: Supplementary Variables",
    "Data Sources: Unclassified",
]
CANONICAL_DATA_SOURCE_KEYWORD_SET = set(CANONICAL_DATA_SOURCE_KEYWORDS)
DATA_SOURCE_SECTION_TO_KEYWORD = {
    "wealth topography": "Data Sources: Wealth Topography",
    "wealth inequality": "Data Sources: Wealth Inequality",
    "wealth inequality trends": "Data Sources: Wealth Inequality",
    "taxes on wealth": "Data Sources: Taxes on Wealth",
    "estate inheritance and gift taxes": "Data Sources: Taxes on Wealth",
    "inheritance trends": "Data Sources: Inheritance Trends",
    "supplementary variables": "Data Sources: Supplementary Variables",
    "unclassified": "Data Sources: Unclassified",
}
DATA_SOURCE_KEYWORD_TO_SECTION = {
    keyword: keyword.replace(DATA_SOURCE_KEYWORD_PREFIX, "", 1)
    for keyword in CANONICAL_DATA_SOURCE_KEYWORDS
}
LEGACY_DATA_SOURCE_KEYWORD_MAP = {
    "data sources: wealth inequality trends": "Data Sources: Wealth Inequality",
    "data sources: estate inheritance and gift taxes": "Data Sources: Taxes on Wealth",
    "data sources: taxes on wealth": "Data Sources: Taxes on Wealth",
    "data sources: wealth topography": "Data Sources: Wealth Topography",
    "data sources: wealth inequality": "Data Sources: Wealth Inequality",
    "data sources: inheritance trends": "Data Sources: Inheritance Trends",
    "data sources: supplementary variables": "Data Sources: Supplementary Variables",
    "data sources: unclassified": "Data Sources: Unclassified",
}
BIB_EXPORT_BLOCKED_FIELDS = {"section"}
DIGITAL_LIBRARY_CONFLICT_FIELDS = ["title", "author", "year", "journal", "doi", "publisher", "abstract"]

DEFAULT_REGISTRY = OrderedDict(
    [
        ("version", 1),
        (
            "config",
            OrderedDict(
                [
                    ("digital_bib_output", DEFAULT_DIGITAL_BIB_PATH),
                    ("bib_output", DEFAULT_DATA_BIB_PATH),
                    ("wealth_bib_input", DEFAULT_WEALTH_BIB_PATH),
                    ("both_bib_output", DEFAULT_BOTH_BIB_PATH),
                    ("wealth_change_log", DEFAULT_WEALTH_CHANGE_LOG_PATH),
                    ("bibbase_profile_source_url", ""),
                    ("bibbase_timeout_seconds", 20),
                    ("online_bib_reference_url", ""),
                    ("online_bib_timeout_seconds", 20),
                    ("wealth_online_bib_reference_url", ""),
                    ("wealth_online_bib_timeout_seconds", 20),
                    ("dictionary_template", DEFAULT_DICTIONARY_PATH),
                    ("dictionary_output", DEFAULT_DICTIONARY_PATH),
                ]
            ),
        ),
        ("records", []),
    ]
)


def now_utc() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def normalize_text(value: str) -> str:
    return normalize_whitespace(value).lower()


def normalize_url(value: str) -> str:
    val = normalize_whitespace(value)
    if not val:
        return ""
    parsed = urlparse(val)
    path = re.sub(r"/+", "/", parsed.path).rstrip("/")
    netloc = parsed.netloc.lower()
    scheme = parsed.scheme.lower() if parsed.scheme else "https"
    rebuilt = f"{scheme}://{netloc}{path}"
    if parsed.query:
        rebuilt = f"{rebuilt}?{parsed.query}"
    if parsed.fragment:
        rebuilt = f"{rebuilt}#{parsed.fragment}"
    return rebuilt


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def sanitize_excel_string(value) -> str:
    if value is None:
        return ""
    text = str(value)
    return "".join(ch for ch in text if _is_valid_excel_xml_char(ord(ch)))


def sanitize_sheet_name(name, fallback: str = "Sheet1") -> str:
    text = INVALID_SHEET_CHARS_RE.sub(" ", sanitize_excel_string(name))
    text = normalize_whitespace(text).strip("'")
    if not text:
        text = INVALID_SHEET_CHARS_RE.sub(" ", sanitize_excel_string(fallback))
        text = normalize_whitespace(text).strip("'")
    return (text or "Sheet1")[:31]


def sanitize_excel_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ""
    return sanitize_excel_string(value)


def _is_valid_excel_xml_char(codepoint: int) -> bool:
    return codepoint in (0x9, 0xA, 0xD) or 0x20 <= codepoint <= 0xD7FF or 0xE000 <= codepoint <= 0xFFFD or 0x10000 <= codepoint <= 0x10FFFF


def _coerce_xlsx_payload(xlsx_source) -> Tuple[bytes, str]:
    if isinstance(xlsx_source, Path):
        return xlsx_source.read_bytes(), str(xlsx_source)
    if isinstance(xlsx_source, str):
        path = Path(xlsx_source)
        return path.read_bytes(), str(path)
    if isinstance(xlsx_source, io.BytesIO):
        return xlsx_source.getvalue(), "<memory>"
    if isinstance(xlsx_source, (bytes, bytearray)):
        return bytes(xlsx_source), "<memory>"
    raise TypeError(f"Unsupported xlsx source type: {type(xlsx_source)!r}")


def _openpyxl_validate_workbook(payload: bytes) -> None:
    load_workbook = _load_openpyxl_workbook()
    workbook = load_workbook(io.BytesIO(payload), read_only=True)
    try:
        _ = workbook.sheetnames
    finally:
        workbook.close()


def _load_openpyxl_workbook():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel validation. Install it with: python3 -m pip install openpyxl"
        ) from exc
    return load_workbook


def workbook_to_xlsx_bytes(workbook) -> bytes:
    workbook_buffer = io.BytesIO()
    workbook.save(workbook_buffer)
    return workbook_buffer.getvalue()


def validate_xlsx_file(xlsx_source) -> None:
    payload, label = _coerce_xlsx_payload(xlsx_source)
    try:
        with zipfile.ZipFile(io.BytesIO(payload), "r") as zf:
            bad_member = zf.testzip()
            if bad_member:
                raise RuntimeError(f"bad zip member: {bad_member}")
            for required in REQUIRED_XLSX_MEMBERS:
                if required not in zf.namelist():
                    raise RuntimeError(f"missing workbook member: {required}")
            sheet_path, _ = locate_sources_sheet(zf)
            ET.fromstring(zf.read(sheet_path))
        _openpyxl_validate_workbook(payload)
    except (zipfile.BadZipFile, KeyError, ET.ParseError, RuntimeError) as exc:
        raise RuntimeError(f"Generated dictionary workbook is invalid ({label}): {exc}") from exc


def load_json_yaml(path: Path) -> dict:
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8").strip()
    if not text:
        return {}
    return json.loads(text)


def dump_json_yaml(path: Path, payload: dict) -> None:
    ensure_parent(path)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=False) + "\n", encoding="utf-8")


def load_registry(path: Path) -> dict:
    data = load_json_yaml(path)
    if not data:
        return deepcopy(DEFAULT_REGISTRY)
    return data


def save_registry(path: Path, data: dict) -> None:
    dump_json_yaml(path, data)


def column_name(idx: int) -> str:
    name = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        name = chr(65 + rem) + name
    return name


def parse_bib_entries(text: str) -> Dict[str, dict]:
    entries: Dict[str, dict] = {}
    i = 0
    while True:
        at = text.find("@", i)
        if at < 0:
            break
        brace = text.find("{", at)
        if brace < 0:
            break
        entry_type = text[at + 1 : brace].strip().lower()
        j = brace + 1
        depth = 1
        while j < len(text):
            ch = text[j]
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    break
            j += 1
        raw = text[brace + 1 : j]
        if "," not in raw:
            i = j + 1
            continue
        key, fields_blob = raw.split(",", 1)
        key = key.strip()
        fields = parse_bib_fields(fields_blob)
        entries[key] = {"entry_type": entry_type, "fields": fields}
        i = j + 1
    return entries


def parse_bib_fields(blob: str) -> OrderedDict:
    fields = OrderedDict()
    p = 0
    n = len(blob)
    while p < n:
        while p < n and blob[p] in " \n\r\t,":
            p += 1
        if p >= n:
            break
        eq = blob.find("=", p)
        if eq < 0:
            break
        name = blob[p:eq].strip().lower()
        p = eq + 1
        while p < n and blob[p] in " \n\r\t":
            p += 1
        if p >= n:
            break

        if blob[p] == "{":
            depth = 1
            p += 1
            start = p
            while p < n and depth > 0:
                if blob[p] == "{":
                    depth += 1
                elif blob[p] == "}":
                    depth -= 1
                p += 1
            value = blob[start : p - 1]
        elif blob[p] == '"':
            p += 1
            start = p
            while p < n and blob[p] != '"':
                if blob[p] == "\\":
                    p += 1
                p += 1
            value = blob[start:p]
            p += 1
        else:
            start = p
            while p < n and blob[p] not in ",\n\r":
                p += 1
            value = blob[start:p].strip()

        fields[name] = value.strip()
        comma = blob.find(",", p)
        if comma < 0:
            break
        p = comma + 1
    return fields


def format_bib_value(value: str) -> str:
    return "{" + value.replace("\n", " ").strip() + "}"


def split_keywords(value: str) -> List[str]:
    tokens: List[str] = []
    seen = set()
    for raw in str(value or "").replace("\n", " ").split(","):
        token = normalize_whitespace(raw)
        token_key = token.lower()
        if token and token_key not in seen:
            tokens.append(token)
            seen.add(token_key)
    return tokens


def format_keywords(tokens: Iterable[str]) -> str:
    return ",".join(split_keywords(",".join(str(token) for token in tokens)))


def is_data_source_keyword(token: str) -> bool:
    return normalize_whitespace(str(token)).startswith(DATA_SOURCE_KEYWORD_PREFIX)


def is_canonical_data_source_keyword(token: str) -> bool:
    return normalize_whitespace(str(token)) in CANONICAL_DATA_SOURCE_KEYWORD_SET


def canonical_data_source_keyword_for_section(section: str) -> str:
    return DATA_SOURCE_SECTION_TO_KEYWORD.get(normalize_text(str(section)), "")


def canonicalize_data_source_keyword(token: str) -> str:
    text = normalize_whitespace(str(token))
    if text in CANONICAL_DATA_SOURCE_KEYWORD_SET:
        return text
    return LEGACY_DATA_SOURCE_KEYWORD_MAP.get(text.lower(), "")


def data_source_keywords_from_value(value: str, *, canonical_only: bool = True) -> List[str]:
    out = []
    for token in split_keywords(value):
        if canonical_only:
            if is_canonical_data_source_keyword(token):
                out.append(token)
        elif is_data_source_keyword(token):
            out.append(token)
    return out


def non_data_source_keywords_from_value(value: str) -> List[str]:
    return [token for token in split_keywords(value) if not is_data_source_keyword(token)]


def set_data_source_keyword(keywords: str, data_source_keyword: str) -> str:
    canonical = canonicalize_data_source_keyword(data_source_keyword)
    if not canonical:
        canonical = data_source_keyword if is_canonical_data_source_keyword(data_source_keyword) else ""
    tokens = non_data_source_keywords_from_value(keywords)
    return format_keywords(([canonical] if canonical else []) + tokens)


def strip_data_source_keywords(keywords: str) -> str:
    return format_keywords(non_data_source_keywords_from_value(keywords))


def migrate_keywords_for_data_source_section(section: str, keywords: str) -> dict:
    before_tokens = split_keywords(keywords)
    before_data_source_keywords = [token for token in before_tokens if is_data_source_keyword(token)]
    non_data_source_tokens = [token for token in before_tokens if not is_data_source_keyword(token)]
    canonical = canonical_data_source_keyword_for_section(section)
    uncertain = bool(normalize_whitespace(str(section)) and not canonical)
    after_tokens = ([canonical] if canonical else []) + non_data_source_tokens
    after_keywords = format_keywords(after_tokens)
    return {
        "before_keywords": format_keywords(before_tokens),
        "after_keywords": after_keywords,
        "before_data_source_keywords": before_data_source_keywords,
        "after_data_source_keywords": data_source_keywords_from_value(after_keywords),
        "backfilled": bool(canonical and not normalize_whitespace(str(keywords))),
        "changed": format_keywords(before_tokens) != after_keywords,
        "uncertain": uncertain,
        "canonical_keyword": canonical,
    }


def sanitize_bib_fields_for_export(fields: dict) -> OrderedDict:
    out = OrderedDict()
    for raw_name, raw_value in (fields or {}).items():
        name = normalize_whitespace(str(raw_name)).lower()
        value = normalize_whitespace(str(raw_value))
        if not name or not value or name in BIB_EXPORT_BLOCKED_FIELDS:
            continue
        if name == "keywords":
            tokens = []
            for token in split_keywords(value):
                if is_data_source_keyword(token) and not is_canonical_data_source_keyword(token):
                    continue
                tokens.append(token)
            value = format_keywords(tokens)
            if not value:
                continue
        out[name] = value
    return out


def parsed_bib_entry_from_record(record: dict) -> dict:
    bib = record.get("bib", {}) or {}
    fields = OrderedDict()
    for field_name in BIB_FIELD_ORDER:
        value = normalize_whitespace(str(bib.get(field_name, "")))
        if value:
            fields[field_name] = value
    extras = bib.get("extra_fields", {}) or {}
    for field_name in sorted(extras.keys()):
        name = normalize_whitespace(str(field_name)).lower()
        value = normalize_whitespace(str(extras[field_name]))
        if name and value and name not in fields:
            fields[name] = value
    return {
        "entry_type": normalize_whitespace(str(bib.get("entry_type", "misc"))).lower() or "misc",
        "fields": sanitize_bib_fields_for_export(fields),
    }


def record_is_data_source(record: dict) -> bool:
    return len(data_source_keywords_from_value((record.get("bib", {}) or {}).get("keywords", ""))) == 1


def _normalized_conflict_value(field_name: str, value: str) -> str:
    if field_name in {"year", "doi"}:
        return normalize_whitespace(str(value)).lower()
    return normalize_text(str(value))


def _is_url_like_field(field_name: str) -> bool:
    name = normalize_whitespace(str(field_name)).lower()
    return (
        name == "url"
        or name.startswith("url_")
        or "file" in name
        or "appendix" in name
        or "replication" in name
        or "codebook" in name
        or "data" in name
    )


def _append_preserved_field(fields: OrderedDict, field_name: str, value: str) -> str:
    if not normalize_whitespace(value):
        return ""
    existing_values = {normalize_whitespace(str(v)) for v in fields.values()}
    if normalize_whitespace(value) in existing_values:
        return ""
    base = normalize_whitespace(field_name).lower()
    idx = 2
    while f"{base}_{idx}" in fields:
        idx += 1
    new_name = f"{base}_{idx}"
    fields[new_name] = value
    return new_name


def _merge_keyword_values(existing_value: str, incoming_value: str) -> str:
    return format_keywords(split_keywords(existing_value) + split_keywords(incoming_value))


def _entry_source_label(source: dict) -> str:
    if source.get("record_id"):
        return normalize_whitespace(str(source.get("record_id", "")))
    if source.get("kind") == "wealth_research":
        return f"wealth:{source.get('key', '')}"
    return normalize_whitespace(str(source.get("key", "")))


def build_digital_library_entries(records: List[dict], wealth_entries: Dict[str, dict]) -> Tuple[Dict[str, dict], dict]:
    items = []
    for key, entry in sorted((wealth_entries or {}).items(), key=lambda item: item[0].lower()):
        clean_entry = {
            "entry_type": normalize_whitespace(str(entry.get("entry_type", "misc"))).lower() or "misc",
            "fields": sanitize_bib_fields_for_export(entry.get("fields", {}) or {}),
        }
        items.append(
            {
                "key": normalize_whitespace(key),
                "entry": clean_entry,
                "kind": "wealth_research",
                "record_id": "",
                "priority": 0,
            }
        )

    for record in records:
        key = normalize_whitespace(record.get("citekey", "")) or normalize_whitespace(record.get("source", ""))
        if not key:
            continue
        is_data_source = record_is_data_source(record)
        items.append(
            {
                "key": key,
                "entry": parsed_bib_entry_from_record(record),
                "kind": "data_source" if is_data_source else "registry",
                "record_id": normalize_whitespace(str(record.get("id", ""))),
                "priority": 2 if is_data_source else 1,
            }
        )

    merged: Dict[str, dict] = {}
    source_meta: Dict[str, List[dict]] = {}
    report = {
        "duplicate_citekeys": [],
        "keyword_unions": [],
        "multi_data_source_keyword_exports": [],
        "bibliographic_conflicts": [],
        "preserved_url_like_fields": [],
        "dropped_data_source_keywords": [],
    }
    duplicate_report_by_key = {}

    for item in items:
        key = item["key"]
        if not key:
            continue
        entry = item["entry"]
        dropped = []
        original_keywords = normalize_whitespace(str((entry.get("fields", {}) or {}).get("keywords", "")))
        for token in split_keywords(original_keywords):
            if is_data_source_keyword(token) and not is_canonical_data_source_keyword(token):
                dropped.append(token)
        if dropped:
            report["dropped_data_source_keywords"].append(
                {"citekey": key, "source": _entry_source_label(item), "keywords": dropped}
            )

        if key not in merged:
            merged[key] = {
                "entry_type": entry.get("entry_type", "misc"),
                "fields": OrderedDict(entry.get("fields", {}) or {}),
                "_priority": item["priority"],
            }
            source_meta[key] = [item]
            continue

        if len(source_meta[key]) == 1:
            duplicate_report_by_key[key] = {"citekey": key, "sources": [_entry_source_label(source_meta[key][0])]}
            report["duplicate_citekeys"].append(duplicate_report_by_key[key])
        duplicate_report_by_key[key]["sources"].append(_entry_source_label(item))

        current = merged[key]
        fields = current["fields"]
        incoming_fields = entry.get("fields", {}) or {}
        existing_priority = current.get("_priority", 0)
        if item["priority"] > existing_priority:
            current["entry_type"] = entry.get("entry_type", current.get("entry_type", "misc"))
            current["_priority"] = item["priority"]

        before_keywords = normalize_whitespace(str(fields.get("keywords", "")))
        incoming_keywords = normalize_whitespace(str(incoming_fields.get("keywords", "")))
        if incoming_keywords:
            fields["keywords"] = _merge_keyword_values(before_keywords, incoming_keywords)
            if normalize_whitespace(str(fields.get("keywords", ""))) != before_keywords:
                report["keyword_unions"].append(
                    {
                        "citekey": key,
                        "source": _entry_source_label(item),
                        "keywords": fields.get("keywords", ""),
                    }
                )

        for field_name, incoming_value in incoming_fields.items():
            if field_name == "keywords" or field_name in BIB_EXPORT_BLOCKED_FIELDS:
                continue
            existing_value = normalize_whitespace(str(fields.get(field_name, "")))
            if not existing_value:
                fields[field_name] = incoming_value
                continue
            if _normalized_conflict_value(field_name, existing_value) == _normalized_conflict_value(field_name, incoming_value):
                continue
            if field_name in DIGITAL_LIBRARY_CONFLICT_FIELDS:
                choose_incoming = item["priority"] > existing_priority
                report["bibliographic_conflicts"].append(
                    {
                        "citekey": key,
                        "field": field_name,
                        "kept": incoming_value if choose_incoming else existing_value,
                        "other": existing_value if choose_incoming else incoming_value,
                        "incoming_source": _entry_source_label(item),
                    }
                )
                if choose_incoming:
                    fields[field_name] = incoming_value
                continue
            if _is_url_like_field(field_name):
                if item["priority"] > existing_priority:
                    old_value = existing_value
                    fields[field_name] = incoming_value
                    preserved_field = _append_preserved_field(fields, field_name, old_value)
                    if preserved_field:
                        report["preserved_url_like_fields"].append(
                            {
                                "citekey": key,
                                "field": field_name,
                                "preserved_as": preserved_field,
                                "source": "previous",
                            }
                        )
                else:
                    preserved_field = _append_preserved_field(fields, field_name, incoming_value)
                    if preserved_field:
                        report["preserved_url_like_fields"].append(
                            {
                                "citekey": key,
                                "field": field_name,
                                "preserved_as": preserved_field,
                                "source": _entry_source_label(item),
                            }
                        )
                continue
        source_meta[key].append(item)

    for key, entry in merged.items():
        fields = entry.get("fields", {}) or {}
        ds_keywords = data_source_keywords_from_value(fields.get("keywords", ""))
        if len(ds_keywords) > 1:
            report["multi_data_source_keyword_exports"].append(
                {
                    "citekey": key,
                    "keywords": ds_keywords,
                    "sources": [_entry_source_label(source) for source in source_meta.get(key, [])],
                }
            )
        entry.pop("_priority", None)

    return merged, report


def render_bib_entry(key: str, record: dict) -> str:
    bib = record.get("bib", {})
    entry_type = bib.get("entry_type", "misc").strip().lower() or "misc"
    ordered_fields = [
        "title",
        "author",
        "year",
        "month",
        "journal",
        "booktitle",
        "volume",
        "number",
        "pages",
        "institution",
        "publisher",
        "doi",
        "url",
        "urldate",
        "abstract",
        "keywords",
        "note",
    ]

    fields = OrderedDict()
    for f in ordered_fields:
        if f in BIB_EXPORT_BLOCKED_FIELDS:
            continue
        val = bib.get(f, "")
        if normalize_whitespace(str(val)):
            fields[f] = str(val)

    extras = bib.get("extra_fields", {}) or {}
    for k in sorted(extras.keys()):
        if normalize_whitespace(str(k)).lower() in BIB_EXPORT_BLOCKED_FIELDS:
            continue
        val = normalize_whitespace(str(extras[k]))
        if val:
            fields[k] = val

    lines = [f"@{entry_type}{{{key},"]
    last_index = len(fields) - 1
    for idx, (name, value) in enumerate(fields.items()):
        tail = "," if idx != last_index else ""
        lines.append(f"  {name} = {format_bib_value(value)}{tail}")
    lines.append("}")
    return "\n".join(lines)


def render_parsed_bib_entry(key: str, entry: dict, field_order: List[str] = None) -> str:
    order = field_order or BIB_FIELD_ORDER
    entry_type = normalize_whitespace(str(entry.get("entry_type", "misc"))).lower() or "misc"
    source_fields = entry.get("fields", {}) or {}

    ordered_fields = OrderedDict()
    for field_name in order:
        if normalize_whitespace(str(field_name)).lower() in BIB_EXPORT_BLOCKED_FIELDS:
            continue
        value = normalize_whitespace(str(source_fields.get(field_name, "")))
        if value:
            ordered_fields[field_name] = value

    for field_name in sorted(source_fields.keys()):
        if normalize_whitespace(str(field_name)).lower() in BIB_EXPORT_BLOCKED_FIELDS:
            continue
        if field_name in ordered_fields or field_name in order:
            continue
        value = normalize_whitespace(str(source_fields.get(field_name, "")))
        if value:
            ordered_fields[field_name] = value

    lines = [f"@{entry_type}{{{key},"]
    last_idx = len(ordered_fields) - 1
    for idx, (name, value) in enumerate(ordered_fields.items()):
        tail = "," if idx != last_idx else ""
        lines.append(f"  {name} = {format_bib_value(value)}{tail}")
    lines.append("}")
    return "\n".join(lines)


def write_parsed_bib_entries(path: Path, entries: Dict[str, dict], field_order: List[str] = None) -> None:
    rendered = []
    for key in sorted(entries.keys(), key=lambda k: k.lower()):
        rendered.append(render_parsed_bib_entry(key, entries[key], field_order=field_order))
    ensure_parent(path)
    path.write_text("\n\n".join(rendered).strip() + "\n", encoding="utf-8")


def records_sorted(records: List[dict]) -> List[dict]:
    return sorted(records, key=lambda r: (normalize_text(r.get("source", "")), normalize_text(r.get("citekey", ""))))


def get_cell_value(cell: ET.Element, shared: List[str], ns: Dict[str, str]) -> str:
    ctype = cell.attrib.get("t")
    if ctype == "inlineStr":
        node = cell.find("a:is/a:t", ns)
        return node.text if node is not None and node.text else ""
    v = cell.find("a:v", ns)
    if v is None:
        return ""
    raw = v.text or ""
    if ctype == "s" and raw.isdigit():
        idx = int(raw)
        return shared[idx] if idx < len(shared) else ""
    return raw


def locate_sources_sheet(zf: zipfile.ZipFile) -> Tuple[str, str]:
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_map = {r.attrib["Id"]: r.attrib["Target"] for r in rels.findall(f"{{{NS_PKG_REL}}}Relationship")}
    ns = {"a": NS_MAIN, "r": NS_REL}
    expected_sheet_name = sanitize_sheet_name(EXPECTED_SOURCES_SHEET_NAME)
    for sheet in workbook.findall("a:sheets/a:sheet", ns):
        if sanitize_sheet_name(sheet.attrib.get("name", "")) == expected_sheet_name:
            rid = sheet.attrib.get(f"{{{NS_REL}}}id")
            if rid:
                target = rid_map[rid]
                if target.startswith("/"):
                    target = target.lstrip("/")
                elif not target.startswith("xl/"):
                    target = "xl/" + target
                return target, sheet.attrib.get("sheetId", "")
    raise RuntimeError("Sources sheet not found in workbook")


def read_sources_sheet(xlsx_path: Path) -> List[dict]:
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        sheet_path, _ = locate_sources_sheet(zf)
        shared = []
        ns = {"a": NS_MAIN}
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", ns):
                text = "".join((t.text or "") for t in si.findall(".//a:t", ns))
                shared.append(text)

        ws = ET.fromstring(zf.read(sheet_path))
        sheet_data = ws.find("a:sheetData", ns)
        if sheet_data is None:
            return []

        headers = {}
        out = []
        for row in sheet_data.findall("a:row", ns):
            ridx = int(row.attrib.get("r", "0"))
            vals = {}
            for cell in row.findall("a:c", ns):
                ref = cell.attrib.get("r", "")
                col = re.match(r"[A-Z]+", ref)
                if not col:
                    continue
                vals[col.group(0)] = get_cell_value(cell, shared, ns)
            if ridx == 1:
                headers = vals
                continue

            if not headers:
                continue

            record = {}
            any_value = False
            for idx in range(1, len(SOURCES_HEADERS) + 1):
                col = column_name(idx)
                header = headers.get(col, SOURCES_HEADERS[idx - 1])
                val = vals.get(col, "")
                record[header] = val
                if normalize_whitespace(val):
                    any_value = True
            if any_value:
                out.append(record)
    return out


def xml_cell(col_idx: int, row_idx: int, value: str) -> str:
    col = column_name(col_idx)
    ref = f"{col}{row_idx}"
    escaped = escape(sanitize_excel_value(value))
    return f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{escaped}</t></is></c>'


def build_sources_sheet_xml(rows: List[dict]) -> bytes:
    header = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    body = [
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
    ]
    max_row = len(rows) + 1
    body.append(f'<dimension ref="A1:R{max_row}"/>')
    body.append('<sheetViews><sheetView workbookViewId="0"/></sheetViews>')
    body.append('<sheetFormatPr defaultRowHeight="15"/>')
    body.append('<sheetData>')

    all_rows = [OrderedDict((h, h) for h in SOURCES_HEADERS)]
    for row in rows:
        ordered = OrderedDict()
        for h in SOURCES_HEADERS:
            ordered[h] = row.get(h, "")
        all_rows.append(ordered)

    for r_idx, row in enumerate(all_rows, start=1):
        body.append(f'<row r="{r_idx}">')
        for c_idx, h in enumerate(SOURCES_HEADERS, start=1):
            body.append(xml_cell(c_idx, r_idx, row.get(h, "")))
        body.append('</row>')

    body.append('</sheetData>')
    body.append(f'<autoFilter ref="A1:R{max_row}"/>')
    body.append('</worksheet>')
    return (header + "".join(body)).encode("utf-8")


def validate_xlsx_for_replace(xlsx_path: Path) -> None:
    validate_xlsx_file(xlsx_path)


def get_sources_worksheet(workbook):
    expected_sheet_name = sanitize_sheet_name(EXPECTED_SOURCES_SHEET_NAME)
    for worksheet in workbook.worksheets:
        if sanitize_sheet_name(worksheet.title) == expected_sheet_name:
            return worksheet
    raise RuntimeError("Sources sheet not found in workbook")


def populate_sources_worksheet(worksheet, rows: List[dict]) -> None:
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    for c_idx, header in enumerate(SOURCES_HEADERS, start=1):
        worksheet.cell(row=1, column=c_idx, value=sanitize_excel_string(header))

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(SOURCES_HEADERS, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=sanitize_excel_value(row.get(header, "")))

    max_row = len(rows) + 1
    worksheet.auto_filter.ref = f"A1:R{max_row}"
    if worksheet.freeze_panes == "A1":
        worksheet.freeze_panes = None


def write_xlsx_atomic(output_xlsx: Path, payload: bytes) -> None:
    ensure_parent(output_xlsx)
    tmp_output = None
    try:
        with tempfile.NamedTemporaryFile(
            "wb",
            delete=False,
            dir=str(output_xlsx.parent),
            prefix=f".{output_xlsx.stem}.",
            suffix=output_xlsx.suffix,
        ) as handle:
            handle.write(payload)
            tmp_output = Path(handle.name)
        validate_xlsx_file(tmp_output)
        tmp_output.replace(output_xlsx)
    except Exception:
        if tmp_output is not None:
            try:
                tmp_output.unlink()
            except FileNotFoundError:
                pass
        raise


def write_sources_sheet(template_xlsx: Path, output_xlsx: Path, rows: List[dict]) -> None:
    load_workbook = _load_openpyxl_workbook()
    workbook = load_workbook(template_xlsx)
    try:
        worksheet = get_sources_worksheet(workbook)
        populate_sources_worksheet(worksheet, rows)
        workbook_bytes = workbook_to_xlsx_bytes(workbook)
    finally:
        workbook.close()

    validate_xlsx_file(workbook_bytes)
    write_xlsx_atomic(output_xlsx, workbook_bytes)


def normalize_record(raw: dict) -> dict:
    row = {k: raw.get(k, "") for k in CANONICAL_KEYS if k != "bib"}
    row["bib"] = raw.get("bib", {})
    row["source"] = normalize_whitespace(row.get("source", ""))
    row["citekey"] = normalize_whitespace(row.get("citekey", ""))
    row["section"] = normalize_whitespace(row.get("section", ""))
    row["aggsource"] = normalize_whitespace(row.get("aggsource", ""))
    row["legend"] = normalize_whitespace(row.get("legend", ""))
    row["link"] = normalize_whitespace(row.get("link", ""))
    row["ref_link"] = normalize_whitespace(row.get("ref_link", ""))
    return row


def record_to_sources_sheet_row(record: dict) -> dict:
    return {
        "Section": record.get("section", ""),
        "AggSource": record.get("aggsource", ""),
        "Legend": record.get("legend", ""),
        "Source": record.get("source", ""),
        "Data_Type": record.get("data_type", ""),
        "Link": record.get("link", ""),
        "Ref_link": record.get("ref_link", ""),
        "Citekey": record.get("citekey", ""),
        "Inclusion_in_Warehouse": record.get("inclusion_in_warehouse", ""),
        "Multigeo_Reference": record.get("multigeo_reference", ""),
        "Metadata": record.get("metadata", ""),
        "Metadatalink": record.get("metadatalink", ""),
        "QcommentsforTA": record.get("qcommentsforta", ""),
        "TAreply": record.get("tareply", ""),
        "TAcomments": record.get("tacomments", ""),
        "ARJcomments": record.get("arjcomments", ""),
        "ARJreplies": record.get("arjreplies", ""),
        "SeeAggSourcelisthere": record.get("seeaggsourcelisthere", ""),
    }
